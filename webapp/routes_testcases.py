"""用例管理 API"""
import io
import json
from flask import Blueprint, request, jsonify, render_template, send_file
from webapp.db import (
    list_testcases, get_testcase_by_id, create_testcase,
    update_testcase, delete_testcase, get_testcase_by_case_id,
)
from utils.excel_reader import load_testcases

api = Blueprint("testcases_api", __name__)


@api.route("/testcases")
def page_testcases():
    """用例管理页面"""
    return render_template("testcases.html")


@api.route("/api/testcases", methods=["GET"])
def api_list():
    category = request.args.get("category")
    cases = list_testcases(category)
    return jsonify(cases)


@api.route("/api/testcases/<int:id>", methods=["GET"])
def api_get(id):
    case = get_testcase_by_id(id)
    if not case:
        return jsonify({"error": "用例不存在"}), 404
    return jsonify(case)


@api.route("/api/testcases", methods=["POST"])
def api_create():
    data = request.get_json()
    if not data or not data.get("case_id"):
        return jsonify({"error": "case_id 不能为空"}), 400
    existing = get_testcase_by_case_id(data["case_id"])
    if existing:
        return jsonify({"error": f"case_id '{data['case_id']}' 已存在"}), 409
    new_id = create_testcase(data)
    return jsonify({"id": new_id, "case_id": data["case_id"]}), 201


@api.route("/api/testcases/<int:id>", methods=["PUT"])
def api_update(id):
    existing = get_testcase_by_id(id)
    if not existing:
        return jsonify({"error": "用例不存在"}), 404
    data = request.get_json()
    update_testcase(id, data)
    return jsonify({"ok": True})


@api.route("/api/testcases/<int:id>", methods=["DELETE"])
def api_delete(id):
    existing = get_testcase_by_id(id)
    if not existing:
        return jsonify({"error": "用例不存在"}), 404
    delete_testcase(id)
    return jsonify({"ok": True})


@api.route("/api/testcases/import", methods=["POST"])
def api_import_excel():
    """从 testdata/ 目录下的 Excel 文件批量导入用例"""
    data = request.get_json() or {}
    filename = data.get("filename", "login_cases.xlsx")

    try:
        cases = load_testcases(filename)
    except FileNotFoundError:
        return jsonify({"error": f"文件 '{filename}' 不存在于 testdata/ 目录"}), 404

    imported = 0
    skipped = 0
    for case in cases:
        existing = get_testcase_by_case_id(case["case_id"])
        if existing:
            skipped += 1
            continue
        # Excel reader 返回的 body 是 dict、expected_contains 是 list，需转回字符串
        expected_contains = case.get("expected_contains", "")
        if isinstance(expected_contains, list):
            expected_contains = ";;".join(expected_contains)

        create_testcase({
            "case_id": case["case_id"],
            "description": case.get("description", ""),
            "method": case.get("method", "GET"),
            "path": case.get("path", ""),
            "headers": json.dumps(case.get("headers", {}), ensure_ascii=False) if case.get("headers") else "",
            "body": json.dumps(case.get("body"), ensure_ascii=False) if case.get("body") else "",
            "expected_status": case.get("expected_status", 200),
            "expected_contains": expected_contains,
            "category": case.get("category", "smoke"),
            "extract_path": case.get("extract_path", ""),
            "save_as": case.get("save_as", ""),
        })
        imported += 1

    return jsonify({"imported": imported, "skipped": skipped})


@api.route("/api/testcases/template", methods=["GET"])
def api_download_template():
    """下载 Excel 模版（含示例数据）"""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "测试用例"

    # 表头
    headers = [
        "case_id", "description", "method", "path", "headers", "body",
        "expected_status", "expected_contains", "category",
        "extract_path", "save_as",
    ]

    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color="2c3e50", end_color="2c3e50", fill_type="solid")
    header_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin"),
    )

    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
        cell.border = thin_border

    # 示例数据
    examples = [
        {
            "case_id": "TC001",
            "description": "正常登录-获取Token",
            "method": "POST",
            "path": "/auth/login",
            "headers": "",
            "body": '{"username":"emilys","password":"emilyspass"}',
            "expected_status": 200,
            "expected_contains": "accessToken;;refreshToken",
            "category": "smoke",
            "extract_path": "accessToken",
            "save_as": "token",
        },
        {
            "case_id": "TC002",
            "description": "用token获取当前用户信息",
            "method": "GET",
            "path": "/auth/me",
            "headers": '{"Authorization":"Bearer {{token}}"}',
            "body": "",
            "expected_status": 200,
            "expected_contains": "username;;email",
            "category": "smoke",
            "extract_path": "",
            "save_as": "",
        },
        {
            "case_id": "TC003",
            "description": "获取用户列表",
            "method": "GET",
            "path": "/users",
            "headers": "",
            "body": "",
            "expected_status": 200,
            "expected_contains": "users",
            "category": "positive",
            "extract_path": "",
            "save_as": "",
        },
        {
            "case_id": "TC004",
            "description": "错误密码登录",
            "method": "POST",
            "path": "/auth/login",
            "headers": "",
            "body": '{"username":"emilys","password":"wrongpass"}',
            "expected_status": 400,
            "expected_contains": "Invalid credentials",
            "category": "negative",
            "extract_path": "",
            "save_as": "",
        },
        {
            "case_id": "TC005",
            "description": "验证用户数据结构",
            "method": "GET",
            "path": "/users/1",
            "headers": "",
            "body": "",
            "expected_status": 200,
            "expected_contains": "id;;firstName;;lastName;;email;;username",
            "category": "structure",
            "extract_path": "",
            "save_as": "",
        },
    ]

    example_font = Font(size=11)
    example_align = Alignment(vertical="center", wrap_text=True)

    for row_idx, ex in enumerate(examples, 2):
        for col_idx, key in enumerate(headers, 1):
            value = ex.get(key, "")
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = example_font
            cell.alignment = example_align
            cell.border = thin_border

    # 列宽
    col_widths = [10, 24, 8, 24, 30, 40, 14, 28, 12, 14, 10]
    for col_idx, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = w

    # 行高
    ws.row_dimensions[1].height = 24
    for r in range(2, len(examples) + 2):
        ws.row_dimensions[r].height = 20

    # 添加使用说明 sheet
    ws2 = wb.create_sheet("使用说明")
    instructions = [
        ["字段说明"],
        ["case_id", "用例编号，必须唯一，如 TC001"],
        ["description", "用例描述"],
        ["method", "HTTP 方法：GET / POST / PUT / DELETE"],
        ["path", "接口路径，如 /auth/login，支持 {{变量}} 占位符"],
        ["headers", "请求头，JSON 格式，支持 {{变量}} 占位符"],
        ["body", "请求体，JSON 格式，支持 {{变量}} 占位符"],
        ["expected_status", "预期 HTTP 状态码，如 200"],
        ["expected_contains", "期望响应中包含的关键字，多个以 ;; 分隔（AND 关系）"],
        ["category", "用例分类：smoke(冒烟) / positive(正向) / negative(异常) / structure(结构校验)"],
        ["extract_path", "从响应 JSON 中提取字段的路径，如 accessToken / data.id"],
        ["save_as", "提取后保存的变量名，后续用例用 {{变量名}} 引用"],
        [],
        ["变量传递示例（链式调用）"],
        ["TC001 登录时 extract_path=accessToken, save_as=token"],
        ["TC002 的 headers 中填写 {\"Authorization\":\"Bearer {{token}}\"}，执行时会自动替换"],
    ]
    for row_idx, row_data in enumerate(instructions, 1):
        for col_idx, val in enumerate(row_data, 1):
            cell = ws2.cell(row=row_idx, column=col_idx, value=val)
            if row_idx == 1:
                cell.font = Font(bold=True, size=13, color="2c3e50")

    ws2.column_dimensions["A"].width = 22
    ws2.column_dimensions["B"].width = 65

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    return send_file(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        as_attachment=True,
        download_name="testcase_template.xlsx",
    )
