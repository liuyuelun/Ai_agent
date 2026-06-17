"""
从 Excel 文件读取测试用例。

Excel 表头定义：
    case_id | description | method | path | headers | body
    | expected_status | expected_contains | category
    | extract_path | save_as

- headers:    JSON 字符串，支持 {{变量名}} 占位符
- body:       JSON 字符串，支持 {{变量名}} 占位符
- path:       接口路径，支持 {{变量名}} 占位符
- expected_contains: 期望响应中包含的字符串（多个以 ;; 分隔表示 AND 关系）
- extract_path: JSON dot-notation 提取路径，如 accessToken / data.token / users.0.name
- save_as:     提取后保存的变量名，供后续用例通过 {{变量名}} 引用
"""

import json
import os
from typing import List, Dict, Any
import openpyxl

_EXCEL_DIR = os.path.join(os.path.dirname(os.path.dirname(__file__)), "testdata")


def load_testcases(filename: str, sheet_name: str = None) -> List[Dict[str, Any]]:
    """
    读取 Excel 中的测试用例，返回 pytest parametrize 可用的列表。
    """
    filepath = os.path.join(_EXCEL_DIR, filename)
    if not os.path.exists(filepath):
        raise FileNotFoundError(f"Excel 文件不存在: {filepath}")

    wb = openpyxl.load_workbook(filepath)
    ws = wb[sheet_name] if sheet_name else wb.active

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    headers_row = [cell.value for cell in ws[1]]

    cases = []
    for row in rows:
        if not row[0]:
            continue

        case = dict(zip(headers_row, row))

        case["expected_status"] = int(case.get("expected_status", 200))
        case["method"] = (case.get("method") or "GET").upper()

        raw_headers = case.get("headers")
        case["headers"] = json.loads(raw_headers) if raw_headers else {}

        raw_body = case.get("body")
        case["body"] = json.loads(raw_body) if raw_body else None

        raw_contains = case.get("expected_contains") or ""
        case["expected_contains"] = [s.strip() for s in raw_contains.split(";;") if s.strip()]

        # 新增：变量提取字段
        case["extract_path"] = (case.get("extract_path") or "").strip()
        case["save_as"] = (case.get("save_as") or "").strip()

        cases.append(case)

    wb.close()
    return cases
